import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from io import BytesIO

# Function to compute Excel date serial
def excel_date_serial(dt):
    base = datetime(1899, 12, 30)
    return (dt - base).days

# Function to compute local time stamp (fractional day)
def local_time_fraction(time_str):
    h, m, s = map(int, time_str.split(':'))
    return (h / 24) + (m / 1440) + (s / 86400)

# Function to process a single CSV and return daily data
def process_csv(df_csv):
    df_csv['Datetime'] = pd.to_datetime(df_csv['Date'] + ' ' + df_csv['Time'])
    df_csv['Day_Serial'] = df_csv['Datetime'].apply(excel_date_serial)
    df_csv['Local_Time_Stamp'] = df_csv['Time'].apply(local_time_fraction)
    df_csv['Active_Power'] = df_csv['PSum']  # Assuming 'PSum' column from raw data
    df_csv['kW'] = -df_csv['Active_Power'] / 1000
    # Group by day
    daily_groups = df_csv.groupby('Day_Serial')
    daily_dfs = {}
    for day_serial, group in daily_groups:
        group = group.sort_values('Local_Time_Stamp')
        daily_dfs[day_serial] = group[['Local_Time_Stamp', 'Active_Power', 'kW']]
    return daily_dfs, df_csv.groupby('Day_Serial')['kW'].agg(['mean', 'max', 'sum']).reset_index()

# Function to build MSB sheet in wide format
def build_msb_sheet(ws, daily_dfs, msb_name):
    # Headers
    ws.cell(row=1, column=1, value="UTC Offset (minutes)")
    col = 2
    days = sorted(daily_dfs.keys())
    max_entries = max(len(daily_dfs[day]) for day in days) if days else 0
    for day in days:
        ws.cell(row=1, column=col, value="Local Time Stamp")
        ws.cell(row=1, column=col+1, value="Active Power (W)")
        ws.cell(row=1, column=col+2, value="kW")
        ws.cell(row=1, column=col+3, value="")  # Empty
        col += 4
    # Data
    if days:
        ws.cell(row=2, column=1, value=days[0])  # First day serial in A2
    for r in range(max_entries + 1):  # +1 for potential header row shift
        col = 2
        for i, day in enumerate(days):
            if r == 0:
                # First data row: day serial in Local Time Stamp if not first block
                if i > 0:
                    ws.cell(row=2 + r, column=col, value=day)
            if r < len(daily_dfs[day]):
                entry = daily_dfs[day].iloc[r]
                ws.cell(row=2 + r, column=col, value=entry['Local_Time_Stamp'])
                ws.cell(row=2 + r, column=col+1, value=entry['Active_Power'])
                ws.cell(row=2 + r, column=col+2, value=entry['kW'])
                ws.cell(row=2 + r, column=col+3, value="")  # Empty
            col += 4

# Function to build Total MSB sheet
def build_total_sheet(ws, daily_summaries):
    ws.cell(row=1, column=1, value="Date")
    ws.cell(row=1, column=2, value="MSB-1")
    ws.cell(row=1, column=3, value="MSB-2")
    ws.cell(row=1, column=4, value="MSB-3")
    ws.cell(row=1, column=5, value="Total Building Load")
    row = 2
    days = sorted(set(daily_summaries[0]['Day_Serial']) | set(daily_summaries[1]['Day_Serial']) | set(daily_summaries[2]['Day_Serial']))
    for day in days:
        kws = [summ[summ['Day_Serial'] == day]['mean'].values[0] if day in summ['Day_Serial'].values else 0 for summ in daily_summaries]
        total = sum(kws)
        ws.cell(row=row, column=1, value=day)
        ws.cell(row=row, column=2, value=kws[0])
        ws.cell(row=row, column=3, value=kws[1])
        ws.cell(row=row, column=4, value=kws[2])
        ws.cell(row=row, column=5, value=total)
        row += 1

# Function to build Load Apportioning sheet (inferred calculations)
def build_load_sheet(ws, daily_summaries):
    # Sample data from your example; adapt as needed
    ws.cell(row=1, column=1, value="Date")
    ws.cell(row=1, column=2, value="Day")
    ws.cell(row=1, column=3, value="MSB No. 2")
    ws.cell(row=1, column=4, value="MSB No. 3")
    ws.cell(row=1, column=5, value="MSB No. 4")
    ws.cell(row=1, column=6, value="MSB No. 5")
    ws.cell(row=1, column=7, value="MSB No. 6")
    ws.cell(row=1, column=8, value="MSB No. 7")
    ws.cell(row=1, column=9, value="EMSB")
    ws.cell(row=1, column=10, value="Maximum Demand, kW")
    row = 3  # Skip row2 as empty in sample
    days = sorted(set(daily_summaries[0]['Day_Serial']) | set(daily_summaries[1]['Day_Serial']) | set(daily_summaries[2]['Day_Serial']))
    for i, day in enumerate(days):
        dt = datetime(1899, 12, 30) + timedelta(days=day)
        day_name = dt.strftime('%A')
        kws_avg = [summ[summ['Day_Serial'] == day]['mean'].values[0] if day in summ['Day_Serial'].values else 0 for summ in daily_summaries]
        max_demand = max(summ[summ['Day_Serial'] == day]['max'].values[0] if day in summ['Day_Serial'].values else 0 for summ in daily_summaries)
        ws.cell(row=row, column=1, value=day)
        ws.cell(row=row, column=2, value=day_name)
        ws.cell(row=row, column=3, value=kws_avg[0])  # MSB1 as No. 2
        ws.cell(row=row, column=4, value=kws_avg[1])  # MSB2 as No. 3
        ws.cell(row=row, column=5, value=kws_avg[2])  # MSB3 as No. 4
        ws.cell(row=row, column=6, value=0)  # Placeholder for No. 5-7
        ws.cell(row=row, column=7, value=0)
        ws.cell(row=row, column=8, value=0)
        ws.cell(row=row, column=9, value=0)  # EMSB placeholder
        ws.cell(row=row, column=10, value=max_demand)
        row += 1
    # Averages, totals, etc.
    ws.cell(row=row, column=2, value="Average")
    for col in range(3, 10):
        ws.cell(row=row, column=col, value=f"=AVERAGE(C3:C{row-1})")  # Formula for average
    row += 1
    ws.cell(row=row, column=2, value="Total")
    for col in range(3, 10):
        ws.cell(row=row, column=col, value=f"=SUM(C3:C{row-2})")
    # Add estimated losses, BEII, etc. (using sample values; adjust formulas/area assumptions)
    row += 2
    ws.cell(row=row, column=1, value="Total Estimated Losses, kW")
    ws.cell(row=row, column=14, value=6.71)  # Sample
    row += 1
    ws.cell(row=row, column=1, value="Building Energy Intensity Index (BEII)")
    # Assume areas: Main Wing 10000 sqm, North 5000, Overall 15000 - replace with real
    total_energy = sum(daily_summaries[i]['sum'].sum() for i in range(3)) / 24  # kWh approximate
    ws.cell(row=row, column=18, value=total_energy / 10000)  # Main Wing
    ws.cell(row=row, column=19, value=total_energy / 5000)  # North Wing
    ws.cell(row=row, column=20, value=total_energy / 15000)  # Overall
    # Add LEII, ACEII, etc. similarly with sample formulas
    row += 1
    ws.cell(row=row, column=1, value="Lighting Energy Intensity Index (LEII)")
    ws.cell(row=row, column=18, value=6.00)  # Sample
    ws.cell(row=row, column=19, value=12.10)
    ws.cell(row=row, column=20, value=8.12)
    # Continue for other sections (AMD Office, etc.) using similar logic/sample data

# Streamlit App
st.title("Energy Analyser Grok")
st.markdown("Upload raw CSV files for MSB 1, 2, 3 to generate the final Excel.")

uploaded_msb1 = st.file_uploader("Raw Data MSB 1.csv", type="csv")
uploaded_msb2 = st.file_uploader("Raw Data MSB 2.csv", type="csv")
uploaded_msb3 = st.file_uploader("Raw Data MSB 3.csv", type="csv")

if uploaded_msb1 and uploaded_msb2 and uploaded_msb3:
    df_msb1 = pd.read_csv(uploaded_msb1, skiprows=2)  # Skip header rows in raw CSV
    df_msb2 = pd.read_csv(uploaded_msb2, skiprows=2)
    df_msb3 = pd.read_csv(uploaded_msb3, skiprows=2)
    
    daily_dfs1, summary1 = process_csv(df_msb1)
    daily_dfs2, summary2 = process_csv(df_msb2)
    daily_dfs3, summary3 = process_csv(df_msb3)
    
    # Create Excel
    output = BytesIO()
    wb = openpyxl.Workbook()
    # MSB 1 sheet
    ws1 = wb.create_sheet("MSB 1")
    build_msb_sheet(ws1, daily_dfs1, "MSB 1")
    # MSB 2
    ws2 = wb.create_sheet("MSB 2")
    build_msb_sheet(ws2, daily_dfs2, "MSB 2")
    # MSB 3
    ws3 = wb.create_sheet("MSB 3")
    build_msb_sheet(ws3, daily_dfs3, "MSB 3")
    # Total MSB
    ws_total = wb.create_sheet("Total MSB")
    build_total_sheet(ws_total, [summary1, summary2, summary3])
    # Load Apportioning
    ws_load = wb.create_sheet("Load Apportioning (Marelli)")
    build_load_sheet(ws_load, [summary1, summary2, summary3])
    # Remove default sheet
    wb.remove(wb['Sheet'])
    wb.save(output)
    
    st.download_button(
        label="Download Final Excel",
        data=output.getvalue(),
        file_name="final_document.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
